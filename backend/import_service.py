import asyncio
import pandas as pd
import io
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from typing import List, Dict, Optional
import models
import re
from moex_service import get_moex_service
from logger import get_logger
from config import settings

log = get_logger("import")


# ── MATH-03 (Sprint 4, Task 5.1): import-hook для MAE/MFE backfill ──
#
# Inline-расчёт MAE/MFE живёт в routers/trades.py (add_trade/close_trade/
# update_trade), но импортированные через /trades/import трейды раньше
# получали NULL mae/mfe — пользователь должен был руками дёргать
# /trades/calculate-mae-mfe POST. Этот hook вызывается из import_trades
# после успешного db.commit() и считает MAE/MFE для свежевставленных
# трейдов в фоне (fire-and-forget), не блокируя HTTP-ответ.

# Strong-ref на background-task'и (Sprint 3 Task 1.1 / PERF-03 pattern):
# без него GC может забрать asyncio.Task до завершения coroutine, и
# loop логирует "Task was destroyed but it is pending". См. middleware.py.
_BG_TASKS: set[asyncio.Task] = set()


async def _backfill_mae_mfe_for_imported_async(
    trade_ids: list[int],
    market_service=None,
    session_factory=None,
) -> None:
    """Background-coroutine: для каждого id в `trade_ids` считает MAE/MFE
    и commit'ит. Любое исключение глотается per-trade — батч не валит
    единичный сбой MOEX.

    Args:
        trade_ids: PK-список свежеимпортированных Trade.
        market_service: DI для тестов; если None — берётся глобальный
            market_data_service из market_service.py.
        session_factory: DI для тестов; если None — database.SessionLocal.
    """
    if not trade_ids:
        return
    if session_factory is None:
        from database import SessionLocal
        session_factory = SessionLocal
    if market_service is None:
        # routers/trades.py создаёт собственный singleton `market_data_service`
        # на module-level — у backend нет глобального синглтона в market_service.py.
        # Инстансируем тут (без I/O в __init__ — см. MarketService).
        from market_service import MarketService
        market_service = MarketService()

    db = session_factory()
    try:
        trades = (
            db.query(models.Trade)
            .filter(models.Trade.id.in_(trade_ids))
            .all()
        )
        for t in trades:
            if t.exit_at is None:
                continue  # open позиции — MAE/MFE неактуальны до закрытия
            if t.mae_price is not None and t.mfe_price is not None:
                continue  # уже посчитано (не пере-затираем)
            try:
                direction = (
                    t.direction.value if hasattr(t.direction, "value") else str(t.direction)
                )
                mae, mfe = await market_service.calculate_mae_mfe(
                    ticker=t.symbol,
                    direction=direction,
                    entry_price=float(t.entry_price),
                    entry_time=t.entry_at,
                    exit_time=t.exit_at,
                    operations=t.operations if t.operations else None,
                    exit_price=float(t.exit_price) if t.exit_price is not None else None,
                )
                if mae is not None and mfe is not None:
                    t.mae_price = Decimal(str(mae))
                    t.mfe_price = Decimal(str(mfe))
            except Exception as exc:
                log.warning("import-hook MAE/MFE failed for trade %s: %s", t.id, exc)
        db.commit()
    except Exception:
        log.exception("import-hook MAE/MFE batch failed")
        try:
            db.rollback()
        except Exception:
            pass
    finally:
        db.close()


def schedule_mae_mfe_backfill(trade_ids: list[int]) -> Optional[asyncio.Task]:
    """Schedule background-task с strong-ref. Вызывать из async-контекста
    (роутер) после commit'а импортированных трейдов.

    В sync-контексте (тесты без работающего loop'а или sync скрипты) —
    no-op: RuntimeError 'no running event loop' гасится.
    Возвращает Task для тестов которые хотят его `await`-ить, либо None.
    """
    if not trade_ids:
        return None
    # ВАЖНО: get_running_loop() ПЕРЕД созданием coroutine — иначе в sync-контексте
    # coroutine создастся и не будет awaited (RuntimeWarning + leak).
    try:
        asyncio.get_running_loop()
    except RuntimeError:
        return None
    task = asyncio.create_task(_backfill_mae_mfe_for_imported_async(list(trade_ids)))
    _BG_TASKS.add(task)
    task.add_done_callback(_BG_TASKS.discard)
    return task

# SEC-07: magic-byte signatures for spreadsheet uploads.
_XLSX_MAGIC = (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")  # ZIP container (xlsx)
_XLS_MAGIC = (b"\xd0\xcf\x11\xe0",)  # OLE2 compound document (legacy xls)


def _validate_spreadsheet_magic(contents: bytes, filename: str) -> None:
    """SEC-07: reject files whose bytes don't match the extension's real format.

    Guards against zip-bombs / disguised payloads fed under .xlsx/.xls names.
    CSV is plain text — skipped.
    """
    name = filename.lower()
    expected: tuple[bytes, ...]
    if name.endswith(".xlsx"):
        expected = _XLSX_MAGIC
    elif name.endswith(".xls"):
        expected = _XLS_MAGIC
    else:
        return
    if not any(contents.startswith(sig) for sig in expected):
        raise ValueError("Файл повреждён или не является Excel/CSV")

# МосБиржа торгуется в МСК (UTC+3). При импорте отчётов с naive-датами
# мы интерпретируем их как МСК и приводим к UTC, чтобы все даты в БД
# были в одной временной шкале (UTC-naive).
MSK_TZ = timezone(timedelta(hours=3))


def _normalize_to_utc_naive(value, assume_tz: timezone = MSK_TZ) -> Optional[datetime]:
    """
    Превращает любое значение даты (str, datetime, pandas.Timestamp) в datetime UTC-naive.

    - Если datetime tz-aware — конвертируем в UTC, отбрасываем tzinfo.
    - Если datetime naive — считаем, что он в `assume_tz` (по умолчанию МСК), и
      переводим в UTC.

    Возвращает None, если значение нельзя распарсить.
    """
    if value is None:
        return None
    try:
        ts = pd.to_datetime(value, errors="coerce")
    except Exception:
        return None
    if pd.isna(ts):
        return None
    dt = ts.to_pydatetime() if hasattr(ts, "to_pydatetime") else ts
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=assume_tz)
    return dt.astimezone(timezone.utc).replace(tzinfo=None)

class TradeManager:
    """
    Менеджер позиций с логикой FIFO.
    
    Логика объединения:
    - Все entry одного тикера в одном направлении до первого exit = ОДНА позиция
    - Добавления (averaging/pyramiding) объединяются в одну позицию
    - При exit закрывается вся позиция целиком
    - Операции сохраняются в массиве operations для аккордеона
    """
    
    def __init__(self):
        self.open_positions = {}  # {symbol: position_dict} - только ОДНА открытая позиция на тикер
        self.completed_trades = []  # Закрытые позиции для сохранения

    def process_trade(self, trade_data: Dict) -> bool:
        """Обработка одной сделки. Возвращает True если сделка обработана, False если пропущена."""
        
        # Валидация входных данных
        quantity = float(trade_data.get('quantity', 0))
        if quantity <= 0:
            # Пропускаем сделки с нулевым или отрицательным количеством
            return False
        
        symbol = trade_data.get('symbol', '').strip()
        if not symbol:
            # Пропускаем сделки без символа
            return False
        
        direction = trade_data.get('direction')
        if direction is None:
            return False
        
        entry_price = float(trade_data.get('entry_price', 0))
        if entry_price <= 0:
            # Пропускаем сделки с нулевой или отрицательной ценой
            return False
        
        entry_at = trade_data.get('entry_at')
        if entry_at is None:
            return False
        
        # Проверяем, есть ли открытая позиция по этому тикеру
        if symbol in self.open_positions:
            open_pos = self.open_positions[symbol]
            
            if open_pos['direction'] != direction:
                # Противоположное направление = ЗАКРЫТИЕ позиции
                self._close_position(trade_data)
            else:
                # Одинаковое направление = ДОБАВЛЕНИЕ к позиции
                self._add_to_position(trade_data)
        else:
            # Нет открытой позиции = новый ВХОД
            self._open_position(trade_data)
        
        return True

    def _open_position(self, trade_data: Dict):
        """Открытие новой позиции"""
        entry_time = trade_data['entry_at']
        
        position = {
            'symbol': trade_data['symbol'],
            'asset_name': trade_data.get('asset_name'),
            'asset_type': trade_data.get('asset_type', 'Stock'),
            'direction': trade_data['direction'],
            'entry_price': float(trade_data['entry_price']),
            'quantity': float(trade_data['quantity']),
            'entry_at': entry_time,
            'exit_at': None,
            'exit_price': None,
            'exit_reason': None,
            'entry_commission': float(trade_data.get('commission', 0)),
            'exit_commission': 0,
            'commission': float(trade_data.get('commission', 0)),
            'swap': float(trade_data.get('swap', 0)),
            'pnl': None,
            'net_pnl': None,
            'holding_time_minutes': None,
            # currency: None означает «использовать дефолт аккаунта», а не подменять «RUB»
            'currency': trade_data.get('currency') or None,
            'notes': trade_data.get('notes'),
            'tags': trade_data.get('tags', []),
            # Для расчёта средневзвешенной цены
            '_total_cost': float(trade_data['entry_price']) * float(trade_data['quantity']),
            '_deal_sum': float(trade_data.get('deal_sum', 0)),
            # Операции для аккордеона
            'operations': [{
                "type": "entry",
                "time": entry_time.strftime("%H:%M:%S") if isinstance(entry_time, datetime) else str(entry_time),
                "date": entry_time.strftime("%d.%m.%Y") if isinstance(entry_time, datetime) else str(entry_time),
                "price": float(trade_data['entry_price']),
                "qty": float(trade_data['quantity']),
                "commission": float(trade_data.get('commission', 0)),
                "direction": trade_data['direction'].value if hasattr(trade_data['direction'], 'value') else str(trade_data['direction'])
            }]
        }
        
        self.open_positions[trade_data['symbol']] = position

    def _add_to_position(self, trade_data: Dict):
        """Добавление к существующей позиции (averaging/pyramiding)"""
        symbol = trade_data['symbol']
        pos = self.open_positions[symbol]
        
        add_qty = float(trade_data['quantity'])
        add_price = float(trade_data['entry_price'])
        add_comm = float(trade_data.get('commission', 0))
        add_swap = float(trade_data.get('swap', 0))
        add_time = trade_data['entry_at']
        
        # Обновляем общий объём и стоимость для средневзвешенной цены
        old_qty = pos['quantity']
        new_qty = old_qty + add_qty
        
        # Средневзвешенная цена (с защитой от деления на ноль)
        pos['_total_cost'] += add_price * add_qty
        if new_qty > 0:
            pos['entry_price'] = pos['_total_cost'] / new_qty
        # Если new_qty == 0, оставляем старую цену (не должно происходить после валидации)
        
        # Суммируем deal_sum для точного расчёта PnL
        pos['_deal_sum'] += float(trade_data.get('deal_sum', 0))
        
        pos['quantity'] = new_qty
        pos['entry_commission'] += add_comm
        pos['commission'] += add_comm
        pos['swap'] += add_swap
        
        # Добавляем операцию в список
        pos['operations'].append({
            "type": "entry",
            "time": add_time.strftime("%H:%M:%S") if isinstance(add_time, datetime) else str(add_time),
            "date": add_time.strftime("%d.%m.%Y") if isinstance(add_time, datetime) else str(add_time),
            "price": add_price,
            "qty": add_qty,
            "commission": add_comm,
            "direction": trade_data['direction'].value if hasattr(trade_data['direction'], 'value') else str(trade_data['direction'])
        })

    def _close_position(self, exit_trade: Dict):
        """Закрытие позиции (полное или частичное)"""
        symbol = exit_trade['symbol']
        pos = self.open_positions[symbol]
        
        exit_qty = float(exit_trade['quantity'])
        exit_price = float(exit_trade['entry_price'])
        exit_comm = float(exit_trade.get('commission', 0))
        exit_swap = float(exit_trade.get('swap', 0))
        exit_time = exit_trade['entry_at']
        exit_deal_sum = float(exit_trade.get('deal_sum', 0))
        
        pos_qty = pos['quantity']
        
        if exit_qty >= pos_qty:
            # Полное закрытие (или переворот позиции)
            close_qty = pos_qty
            remaining_exit_qty = exit_qty - pos_qty
            
            # Пропорция для комиссий
            exit_ratio = close_qty / exit_qty if exit_qty > 0 else 1
            
            pos['exit_at'] = exit_time
            pos['exit_price'] = exit_price
            pos['exit_commission'] = exit_comm * exit_ratio
            pos['commission'] = pos['entry_commission'] + pos['exit_commission']
            pos['swap'] += exit_swap * exit_ratio
            pos['exit_reason'] = 'Manual'
            
            # Добавляем exit операцию
            pos['operations'].append({
                "type": "exit",
                "time": exit_time.strftime("%H:%M:%S") if isinstance(exit_time, datetime) else str(exit_time),
                "date": exit_time.strftime("%d.%m.%Y") if isinstance(exit_time, datetime) else str(exit_time),
                "price": exit_price,
                "qty": close_qty,
                "commission": exit_comm * exit_ratio,
                "direction": exit_trade['direction'].value if hasattr(exit_trade['direction'], 'value') else str(exit_trade['direction'])
            })
            
            # Рассчитываем время удержания
            entry_at = pos['entry_at']
            if entry_at and exit_time:
                delta = exit_time - entry_at
                pos['holding_time_minutes'] = int(delta.total_seconds() / 60)
            
            # Рассчитываем PnL
            entry_deal_sum = pos['_deal_sum']
            exit_deal_sum_part = exit_deal_sum * exit_ratio
            
            if entry_deal_sum and exit_deal_sum_part:
                # Точный расчёт по суммам сделок (уже в рублях)
                if pos['direction'] == models.TradeDirection.LONG:
                    pnl = exit_deal_sum_part - entry_deal_sum
                else:
                    pnl = entry_deal_sum - exit_deal_sum_part
            else:
                # Fallback по ценам
                # Для фьючерсов нужен point_value (цена пункта)
                point_value = Decimal(1)
                moex = get_moex_service()
                if moex.is_futures_ticker(symbol):
                    point_value = moex.get_point_value(symbol)
                
                price_diff = Decimal(str(exit_price)) - Decimal(str(pos['entry_price']))
                if pos['direction'] == models.TradeDirection.SHORT:
                    price_diff = -price_diff
                
                pnl = float(price_diff * Decimal(str(close_qty)) * point_value)
            
            pos['pnl'] = pnl
            pos['net_pnl'] = pnl - pos['commission'] - pos['swap']
            
            # Удаляем временные поля
            pos.pop('_total_cost', None)
            pos.pop('_deal_sum', None)
            
            # Сохраняем закрытую позицию
            self.completed_trades.append(pos)
            
            # Удаляем из открытых
            del self.open_positions[symbol]
            
            # Если остался объём — это переворот позиции
            if remaining_exit_qty > 0:
                flip_trade = exit_trade.copy()
                flip_trade['quantity'] = remaining_exit_qty
                flip_trade['commission'] = exit_comm * (remaining_exit_qty / exit_qty)
                flip_trade['swap'] = exit_swap * (remaining_exit_qty / exit_qty)
                flip_trade['deal_sum'] = exit_deal_sum * (remaining_exit_qty / exit_qty)
                self._open_position(flip_trade)
        
        else:
            # Частичное закрытие (редкий случай)
            close_ratio = exit_qty / pos_qty
            
            # Создаём закрытую часть
            closed_part = pos.copy()
            closed_part['operations'] = list(pos['operations'])  # Deep copy
            closed_part['quantity'] = exit_qty
            closed_part['entry_commission'] = pos['entry_commission'] * close_ratio
            closed_part['exit_commission'] = exit_comm
            closed_part['commission'] = closed_part['entry_commission'] + closed_part['exit_commission']
            closed_part['swap'] = pos['swap'] * close_ratio
            closed_part['exit_at'] = exit_time
            closed_part['exit_price'] = exit_price
            closed_part['exit_reason'] = 'Manual'
            
            # Добавляем exit операцию
            closed_part['operations'].append({
                "type": "exit",
                "time": exit_time.strftime("%H:%M:%S") if isinstance(exit_time, datetime) else str(exit_time),
                "date": exit_time.strftime("%d.%m.%Y") if isinstance(exit_time, datetime) else str(exit_time),
                "price": exit_price,
                "qty": exit_qty,
                "commission": exit_comm,
                "direction": exit_trade['direction'].value if hasattr(exit_trade['direction'], 'value') else str(exit_trade['direction'])
            })
            
            # Время удержания
            if pos['entry_at'] and exit_time:
                delta = exit_time - pos['entry_at']
                closed_part['holding_time_minutes'] = int(delta.total_seconds() / 60)
            
            # PnL для закрытой части
            entry_deal_sum_part = pos['_deal_sum'] * close_ratio
            
            if entry_deal_sum_part and exit_deal_sum:
                # Точный расчёт по суммам сделок (уже в рублях)
                if pos['direction'] == models.TradeDirection.LONG:
                    pnl = exit_deal_sum - entry_deal_sum_part
                else:
                    pnl = entry_deal_sum_part - exit_deal_sum
            else:
                # Fallback по ценам
                # Для фьючерсов нужен point_value (цена пункта)
                point_value = Decimal(1)
                moex = get_moex_service()
                if moex.is_futures_ticker(symbol):
                    point_value = moex.get_point_value(symbol)
                
                price_diff = Decimal(str(exit_price)) - Decimal(str(pos['entry_price']))
                if pos['direction'] == models.TradeDirection.SHORT:
                    price_diff = -price_diff
                
                pnl = float(price_diff * Decimal(str(exit_qty)) * point_value)
            
            closed_part['pnl'] = pnl
            closed_part['net_pnl'] = pnl - closed_part['commission'] - closed_part['swap']
            
            # Удаляем временные поля
            closed_part.pop('_total_cost', None)
            closed_part.pop('_deal_sum', None)
            
            self.completed_trades.append(closed_part)
            
            # Обновляем оставшуюся открытую позицию
            remaining_ratio = 1 - close_ratio
            pos['quantity'] = pos_qty - exit_qty
            pos['_total_cost'] *= remaining_ratio
            pos['_deal_sum'] *= remaining_ratio
            pos['entry_commission'] *= remaining_ratio
            pos['commission'] = pos['entry_commission']
            pos['swap'] *= remaining_ratio

    def finalize(self) -> List[Dict]:
        """Финализация: добавляем оставшиеся открытые позиции"""
        for symbol, pos in self.open_positions.items():
            # Удаляем временные поля
            pos.pop('_total_cost', None)
            pos.pop('_deal_sum', None)
            self.completed_trades.append(pos)
        
        self.open_positions = {}  # Очищаем
        return self.completed_trades


def parse_account_balance(contents: bytes, filename: str = "") -> Dict:
    """
    Парсит информацию о балансе счёта из файла брокера.
    
    Возвращает:
    {
        "initial_balance": float | None,
        "final_balance": float | None,
        "deposits": float | None,
        "withdrawals": float | None,
        "currency": str,
        "period_start": datetime | None,
        "period_end": datetime | None,
    }
    """
    from openpyxl import load_workbook
    
    result = {
        "initial_balance": None,
        "final_balance": None,
        "deposits": None,
        "withdrawals": None,
        "currency": "RUB",
        "period_start": None,
        "period_end": None,
    }
    
    if not filename.lower().endswith(('.xlsx', '.xls')):
        return result
    
    try:
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Ищем информацию в первых 100 строках
        for row in range(1, 100):
            for col in range(1, 20):
                cell_val = ws.cell(row=row, column=col).value
                if not cell_val:
                    continue
                cell_str = str(cell_val).lower()
                
                # Период отчёта
                if 'период' in cell_str or 'period' in cell_str:
                    # Ищем даты в соседних ячейках
                    for c in range(col, min(col + 5, 20)):
                        val = ws.cell(row=row, column=c).value
                        if isinstance(val, datetime):
                            if result["period_start"] is None:
                                result["period_start"] = val
                            else:
                                result["period_end"] = val
                
                # Начальный баланс / Входящий остаток
                if any(x in cell_str for x in ['входящий остаток', 'начальный баланс', 'opening balance', 'на начало периода']):
                    # Ищем число в соседних ячейках
                    for c in range(col, min(col + 10, 30)):
                        val = ws.cell(row=row, column=c).value
                        if val and isinstance(val, (int, float)):
                            result["initial_balance"] = float(val)
                            break
                
                # Конечный баланс / Исходящий остаток
                if any(x in cell_str for x in ['исходящий остаток', 'конечный баланс', 'closing balance', 'на конец периода', 'итого активов']):
                    for c in range(col, min(col + 10, 30)):
                        val = ws.cell(row=row, column=c).value
                        if val and isinstance(val, (int, float)):
                            result["final_balance"] = float(val)
                            break
                
                # Зачисления / Пополнения
                if any(x in cell_str for x in ['зачисление', 'пополнение', 'deposit', 'ввод денег', 'ввод дс']):
                    for c in range(col, min(col + 10, 30)):
                        val = ws.cell(row=row, column=c).value
                        if val and isinstance(val, (int, float)) and float(val) > 0:
                            result["deposits"] = float(val)
                            break
                
                # Списания / Снятия
                if any(x in cell_str for x in ['списание', 'снятие', 'withdrawal', 'вывод денег', 'вывод дс']):
                    for c in range(col, min(col + 10, 30)):
                        val = ws.cell(row=row, column=c).value
                        if val and isinstance(val, (int, float)):
                            result["withdrawals"] = abs(float(val))
                            break
                
                # Валюта
                if 'валюта' in cell_str or 'currency' in cell_str:
                    for c in range(col, min(col + 5, 20)):
                        val = ws.cell(row=row, column=c).value
                        if val and str(val).upper() in ['RUB', 'USD', 'EUR', 'РУБ', 'РУБЛЬ']:
                            result["currency"] = "RUB" if 'РУБ' in str(val).upper() else str(val).upper()
                            break
        
    except Exception as e:
        log.warning(f"Could not parse balance info: {e}")
    
    return result


def parse_margin_fees(contents: bytes) -> Dict[str, float]:
    """
    Парсит раздел 3.4 'Комиссия за маржинальную торговлю' из Excel отчёта Тинькофф.
    Возвращает словарь {дата: сумма} для распределения по позициям.
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    margin_fees = {}
    in_margin_section = False
    
    for row in range(1, 500):
        cell_val = ws.cell(row=row, column=1).value
        
        # Ищем начало раздела 3.4
        if cell_val and '3.4' in str(cell_val) and 'комисси' in str(cell_val).lower():
            in_margin_section = True
            continue
        
        # Конец раздела
        if in_margin_section and cell_val and ('3.5' in str(cell_val) or 'Итого' in str(cell_val)):
            break
        
        if in_margin_section:
            # Колонка 5 = Вид комиссии, 28 = Дата начисления, 44 = Сумма
            fee_type = ws.cell(row=row, column=5).value
            if fee_type and 'MARGIN' in str(fee_type).upper():
                date_val = ws.cell(row=row, column=28).value
                amount = ws.cell(row=row, column=44).value
                
                if date_val and amount:
                    try:
                        date_str = str(date_val)
                        amt = float(amount)
                        if date_str not in margin_fees:
                            margin_fees[date_str] = 0
                        margin_fees[date_str] += amt
                    except (ValueError, TypeError):
                        pass
    
    return margin_fees


def parse_tinkoff_excel(contents: bytes) -> List[Dict]:
    # Parse margin fees first
    margin_fees = parse_margin_fees(contents)
    total_margin_fee = sum(margin_fees.values())
    
    # Read Excel, finding the header row
    df = pd.read_excel(io.BytesIO(contents), header=None)

    # SEC-07: row-cap (DoS / zip-bomb after decompression).
    if len(df) > settings.MAX_IMPORT_ROWS:
        raise ValueError(f"Слишком много строк (макс {settings.MAX_IMPORT_ROWS})")

    start_row = -1
    for i, row in df.iterrows():
        row_str = row.astype(str).str.cat(sep=' ')
        if "Номер сделки" in row_str and "Вид сделки" in row_str:
            start_row = i
            break
            
    if start_row == -1:
        raise ValueError("Could not find trade table in Excel file")
        
    # Reload with correct header
    df = pd.read_excel(io.BytesIO(contents), header=start_row)
    
    # Нормализация колонок: newlines → spaces, схлопываем множественные пробелы
    # Создаём маппинг normalized_name → original_column_name
    cols = { " ".join(str(c).replace('\n', ' ').split()): c for c in df.columns }
    
    def _find_col(*aliases):
        """Ищет колонку по списку алиасов (частичное совпадение). Возвращает оригинальное имя колонки."""
        # Сначала точное совпадение
        for alias in aliases:
            if alias in cols:
                return cols[alias]
        # Затем частичное (алиас как подстрока колонки или наоборот)
        for alias in aliases:
            alias_lower = alias.lower()
            for norm_name, orig_name in cols.items():
                if alias_lower in norm_name.lower() or norm_name.lower() in alias_lower:
                    return orig_name
        return None
    
    # Предварительно находим все нужные колонки один раз (не в цикле!)
    col_date = _find_col('Дата заключения', 'Дата', 'Date')
    col_time = _find_col('Время', 'Time')
    col_side = _find_col('Вид сделки', 'Направление', 'Side')
    col_symbol = _find_col('Код актива', 'Тикер', 'Symbol')
    col_qty = _find_col('Количество', 'Quantity', 'Qty')
    col_deal_sum = _find_col('Сумма сделки', 'Deal Sum', 'Amount')
    col_comm = _find_col('Комиссия брокера', 'Комиссия', 'Commission', 'Fee')
    col_swap = _find_col('Плата за перенос позиций', 'Своп', 'Swap')
    col_price = _find_col('Цена за единицу', 'Цена', 'Price')
    col_name_candidates = ['Сокращенное наименование', 'Наименование', 'Наименование актива']
    col_name = _find_col(*col_name_candidates)
    
    # Валидация: обязательные колонки
    missing = []
    if not col_date: missing.append('Дата')
    if not col_side: missing.append('Вид сделки')
    if not col_symbol: missing.append('Код актива')
    if not col_qty: missing.append('Количество')
    if missing:
        available = list(cols.keys())
        raise ValueError(
            f"Не найдены колонки: {', '.join(missing)}. "
            f"Доступные: {available[:15]}"
        )
    
    manager = TradeManager()
    
    # 1. Extract raw rows
    raw_trades = []
    skipped_rows = 0
    
    for _, row in df.iterrows():
        try:
            date_val = row[col_date] if col_date else None
            time_val = row[col_time] if col_time else None
            side_val = row[col_side] if col_side else None
            symbol_val = row[col_symbol] if col_symbol else None
            qty_val = row[col_qty] if col_qty else None
            deal_sum_val = row[col_deal_sum] if col_deal_sum else None
            
            # Asset Name
            asset_name = None
            if col_name:
                val = row[col_name]
                if pd.notna(val):
                    asset_name = val
            
            if pd.isna(date_val) or pd.isna(symbol_val):
                continue
            
            # Пропускаем повторяющиеся строки-заголовки в середине файла
            if str(date_val).lower() in ['дата заключения', 'дата', 'date']:
                continue
                
            if isinstance(date_val, str):
                date_str = date_val
            elif hasattr(date_val, 'strftime'):
                date_str = date_val.strftime("%d.%m.%Y")
            else:
                continue  # NaT или непонятный тип — пропускаем строку
                
            if time_val is None or (hasattr(time_val, '__class__') and 'NaT' in str(type(time_val))):
                time_str = "00:00:00"
            elif isinstance(time_val, str):
                time_str = time_val
            elif hasattr(time_val, 'strftime'):
                time_str = time_val.strftime("%H:%M:%S")
            else:
                time_str = str(time_val)
            
            # Пропускаем если время - это заголовок
            if str(time_str).lower() in ['время', 'time']:
                continue
                
            try:
                entry_at = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M:%S")
            except ValueError:
                # Пробуем альтернативные форматы даты
                for fmt in ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d.%m.%Y %H:%M"]:
                    try:
                        entry_at = datetime.strptime(f"{date_str} {time_str}", fmt)
                        break
                    except ValueError:
                        continue
                else:
                    log.warning(f"Cannot parse date/time: '{date_str}' / '{time_str}'")
                    continue
            
            side_str = str(side_val).lower()
            if 'репо' in side_str or 'рпс' in side_str:
                continue
                
            if 'покупка' in side_str:
                direction = models.TradeDirection.LONG
            elif 'продажа' in side_str:
                direction = models.TradeDirection.SHORT
            else:
                continue
                
            symbol = str(symbol_val).strip()
            quantity = float(qty_val)
            
            # Commissions
            comm_broker = row.get(col_comm, 0) if col_comm else 0
            
            def parse_comm(val):
                try: return abs(float(val)) if pd.notna(val) else 0.0
                except (ValueError, TypeError): return 0.0
            
            # User correction: Broker commission already includes exchange and clearing fees
            commission = parse_comm(comm_broker)
            
            # Swap / Rollover
            swap_val = row.get(col_swap, 0) if col_swap else 0
            if pd.isna(swap_val) or swap_val == 0:
                swap_val = 0
            swap = parse_comm(swap_val)

            # Deal Sum (for price calc)
            deal_sum = 0.0
            if deal_sum_val is not None and pd.notna(deal_sum_val):
                deal_sum = float(deal_sum_val)
            
            # Fallback price if deal_sum is 0 (unlikely but possible)
            price_per_unit = 0.0
            if col_price:
                pval = row.get(col_price)
                if pval is not None and pd.notna(pval):
                    try:
                        price_per_unit = float(pval)
                    except (ValueError, TypeError):
                        pass

            raw_trades.append({
                "entry_at": entry_at,
                "symbol": symbol,
                "direction": direction,
                "quantity": quantity,
                "deal_sum": deal_sum,
                "commission": commission,
                "swap": swap,
                "asset_name": str(asset_name) if asset_name else None,
                "price_per_unit": price_per_unit
            })

        except Exception as e:
            skipped_rows += 1
            if skipped_rows <= 5:  # Логируем первые 5 ошибок, дальше молчим
                log.warning(f"Error parsing row {row.name}: {e}")
            continue
    
    if skipped_rows > 0:
        log.info(f"Excel parsing: {len(raw_trades)} trades parsed, {skipped_rows} rows skipped")

    # 2. Сортируем сделки по времени для корректной работы FIFO логики
    raw_trades.sort(key=lambda x: x['entry_at'])
    
    # 3. Group by (entry_at, symbol, direction)
    # Since the file is now sorted chronologically, we can iterate and coalesce consecutive matches
    grouped_trades = []
    if raw_trades:
        current_group = raw_trades[0]
        
        for next_trade in raw_trades[1:]:
            if (next_trade['entry_at'] == current_group['entry_at'] and 
                next_trade['symbol'] == current_group['symbol'] and 
                next_trade['direction'] == current_group['direction']):
                
                # Merge
                current_group['quantity'] += next_trade['quantity']
                current_group['deal_sum'] += next_trade['deal_sum']
                current_group['commission'] += next_trade['commission']
                current_group['swap'] += next_trade['swap']
                # Asset name - keep first non-null
                if not current_group['asset_name'] and next_trade['asset_name']:
                    current_group['asset_name'] = next_trade['asset_name']
            else:
                grouped_trades.append(current_group)
                current_group = next_trade
        grouped_trades.append(current_group)

    # ISIN to readable ticker mapping (SPB Exchange uses ISIN codes)
    # For client, it's the same stock - broker decides which exchange to use
    isin_to_ticker = {
        'RU0006944147': 'TATNP',   # Татнефть привилегированные
        'RU0007661625': 'GAZP',    # Газпром
        'RU0009033591': 'TATN',    # Татнефть обыкновенные
        'RU0009062285': 'AFLT',    # Аэрофлот
        'RU0009062467': 'SIBN',    # Газпром нефть
        'RU000A0DKVS5': 'NVTK',    # Новатэк
        'RU000A0JXNU8': 'FLOT',    # Совкомфлот
        'RU000A107UL4': 'T',       # Т-Банк (Т-Технологии)
        'RU000A10ANA1': 'CNRU',    # Циан
        'RU000A0ZZFS9': 'LEAS',    # ЛК Европлан
        'RU000A0JQ9P0': 'RENI',    # Ренессанс Страхование
        'RU000A0JXN21': 'AQUA',    # ИНАРКТИКА
        'RU000A103X66': 'POSI',    # Positive Technologies
        'RU000A106K63': 'VKCO',    # ВК
        'RU000A100K72': 'HNFG',    # Хендерсон
        'RU000A101NJ4': 'SPBE',    # СПБ Биржа
    }

    # Build asset_name -> symbol mapping from non-ISIN entries
    # This helps resolve unknown ISINs by asset name
    asset_name_to_symbol = {}
    for t in grouped_trades:
        sym = t['symbol']
        name = t.get('asset_name')
        # If symbol is not ISIN-like (doesn't start with RU + digits)
        if name and not (sym.startswith('RU') and len(sym) == 12 and sym[2:].isalnum()):
            asset_name_to_symbol[name] = sym
    
    # 3. Process grouped trades through TradeManager
    for t in grouped_trades:
        symbol = t['symbol']
        direction = t['direction']
        quantity = t['quantity']
        deal_sum = t['deal_sum']
        
        # Replace ISIN with readable ticker if known
        if symbol in isin_to_ticker:
            symbol = isin_to_ticker[symbol]
        # If still looks like ISIN, try to resolve by asset_name
        elif symbol.startswith('RU') and len(symbol) == 12 and t.get('asset_name') in asset_name_to_symbol:
            symbol = asset_name_to_symbol[t['asset_name']]
        
        # Infer Asset Type first (needed for price calculation)
        asset_type = "Stock"
        # Futures have format like ETZ5, SiZ4, RIH5 (2-3 letters + month code + year)
        # Используем re.IGNORECASE для учёта тикеров типа SiZ4 (mixed case)
        is_futures = re.search(r'^[A-Za-z]{2,4}[FGHJKMNQUVXZfghjkmnquvxz]\d$', symbol)
        if is_futures:
            asset_type = "Futures"
        # Bonds typically have longer codes or specific patterns
        elif symbol.startswith('SU') or 'ОФЗ' in str(t.get('asset_name', '')):
            asset_type = "Bond"

        # Calculate price based on asset type
        # For futures, use price_per_unit (in native units like USD for commodities)
        # For stocks, use deal_sum / quantity (in RUB)
        if asset_type == "Futures":
            # For futures: use price in points/native currency (e.g., USD for PLD)
            price = t['price_per_unit'] if t['price_per_unit'] else (deal_sum / quantity if quantity else 0)
        else:
            # For stocks: use deal_sum / quantity
            if quantity != 0 and deal_sum != 0:
                price = deal_sum / quantity
            else:
                price = t['price_per_unit']

        trade_data = {
            "symbol": symbol,
            "asset_name": t['asset_name'],
            "asset_type": asset_type,
            "direction": direction,
            "entry_price": price,
            "quantity": quantity,
            "deal_sum": deal_sum, # Pass deal_sum for precise PnL calc
            "entry_at": t['entry_at'],
            "commission": t['commission'],
            "swap": t['swap'],
            "notes": "Imported from Tinkoff Excel",
            "tags": ["Tinkoff", "Imported"]
        }
        
        manager.process_trade(trade_data)
    
    # Финализируем: добавляем оставшиеся открытые позиции
    manager.finalize()
    
    # Распределяем плату за маржинальную торговлю по позициям,
    # которые были открыты в дату начисления комиссии
    for fee_date_str, fee_amount in margin_fees.items():
        if fee_amount <= 0:
            continue
            
        # Парсим дату комиссии
        try:
            fee_date = datetime.strptime(fee_date_str, "%d.%m.%Y").date()
        except (ValueError, TypeError):
            continue
        
        # Находим позиции, которые были перенесены через ночь на эту дату
        # Сначала ищем акции, если нет - берём фьючерсы
        stock_positions = []
        futures_positions = []
        
        for trade in manager.completed_trades:
            entry_date = trade.get('entry_at')
            exit_date = trade.get('exit_at')
            asset_type = trade.get('asset_type', 'Stock')
            
            if entry_date:
                entry_d = entry_date.date() if hasattr(entry_date, 'date') else entry_date
                exit_d = exit_date.date() if exit_date and hasattr(exit_date, 'date') else None
                
                # Плата за плечи берётся только за ПЕРЕНОС через ночь:
                # позиция открыта СТРОГО ДО даты комиссии (entry_d < fee_date)
                # И (не закрыта ИЛИ закрыта в эту дату или позже)
                if entry_d < fee_date and (exit_d is None or exit_d >= fee_date):
                    if asset_type == 'Futures':
                        futures_positions.append(trade)
                    else:
                        stock_positions.append(trade)
        
        # Приоритет: акции, если нет акций - фьючерсы
        positions_on_date = stock_positions if stock_positions else futures_positions
        
        if positions_on_date:
            # Распределяем пропорционально стоимости
            total_value = sum(abs(t.get('entry_price', 0) * t.get('quantity', 0)) for t in positions_on_date)
            if total_value > 0:
                for trade in positions_on_date:
                    position_value = abs(trade.get('entry_price', 0) * trade.get('quantity', 0))
                    trade_fee = fee_amount * (position_value / total_value)
                    trade['swap'] = trade.get('swap', 0) + trade_fee
                    # Пересчитываем net_pnl с учётом нового swap
                    if trade.get('pnl') is not None:
                        trade['net_pnl'] = trade['pnl'] - trade.get('commission', 0) - trade['swap']
    
    # Очищаем временные поля, которые не являются полями модели Trade
    result = []
    fields_to_remove = ['deal_sum', 'price_per_unit', '_total_cost', '_deal_sum']
    for trade in manager.completed_trades:
        clean_trade = {k: v for k, v in trade.items() if k not in fields_to_remove}
        result.append(clean_trade)
    
    # Сортируем по дате входа для корректного отображения в UI
    # Без этой сортировки порядок определяется датой ЗАКРЫТИЯ, что вводит пользователя в заблуждение
    result.sort(key=lambda x: x.get('entry_at') or datetime.min)
    
    return result

def parse_trade_file(contents: bytes, filename: str) -> List[Dict]:
    """
    Парсит файл (CSV/Excel) и возвращает список словарей для создания сделок.
    """
    # SEC-07: validate magic bytes for spreadsheet uploads before any parsing.
    _validate_spreadsheet_magic(contents, filename)

    if filename.endswith(('.xls', '.xlsx')):
        # Try Tinkoff Excel parser first if it looks like a report
        try:
            return parse_tinkoff_excel(contents)
        except ValueError as ve:
            # Row-cap is a hard limit — do not silently fall back to generic.
            if "Слишком много строк" in str(ve):
                raise
            log.debug("Tinkoff Excel parser failed, falling back to generic parser")
        except Exception:
            # Fallback to generic excel parser (implemented below via pandas)
            log.debug("Tinkoff Excel parser failed, falling back to generic parser")

    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        elif filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise ValueError("Unsupported file format")
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Failed to read file: {str(e)}")

    # SEC-07: row-cap (DoS / zip-bomb after decompression).
    if len(df) > settings.MAX_IMPORT_ROWS:
        raise ValueError(f"Слишком много строк (макс {settings.MAX_IMPORT_ROWS})")

    trades = []
    
    # Нормализация имен колонок для Generic парсера
    df.columns = [c.strip() for c in df.columns]
    lower_cols = {c.lower(): c for c in df.columns}


    # Маппинг колонок (можно расширять)
    col_map = {
        'date': ['date', 'time', 'created time', 'date(utc)'],
        'symbol': ['symbol', 'pair', 'instrument', 'contract'],
        'side': ['side', 'type', 'direction', 'operation'],
        'price': ['price', 'avg price', 'entry price', 'avg_price_usd'],
        'quantity': ['amount', 'quantity', 'executed', 'size', 'qty'],
        'pnl': ['pnl', 'realized pnl', 'profit', 'net profit'],
        'fee': ['fee', 'commission'],
        'currency': ['currency', 'ccy', 'валюта', 'валюта расчётов'],
    }

    # Если колонка с датой называется "date(utc)" — данные уже в UTC, иначе считаем МСК.
    # Это спасает от рассинхрона с Binance/Bybit (UTC) vs МосБиржевыми отчётами (МСК).
    def _detect_assume_tz(col_name: str) -> timezone:
        if col_name and "utc" in col_name.lower():
            return timezone.utc
        return MSK_TZ

    def get_col(key):
        for candidate in col_map[key]:
            if candidate in lower_cols:
                return lower_cols[candidate]
        return None

    # Проверяем обязательные поля
    required = ['date', 'symbol', 'side', 'price', 'quantity']
    missing = [key for key in required if get_col(key) is None]
    
    if missing:
        # Если не нашли стандартные, пробуем специфичные для Binance/Bybit
        # Но пока вернем ошибку или пустой список
        log.warning(f"Missing columns: {missing}")
        # Fallback logic could go here
        pass

    date_col = get_col('date')
    symbol_col = get_col('symbol')
    side_col = get_col('side')
    price_col = get_col('price')
    qty_col = get_col('quantity')
    pnl_col = get_col('pnl')
    currency_col = get_col('currency')

    if not all([date_col, symbol_col, side_col, price_col, qty_col]):
         raise ValueError(f"Could not detect required columns. Found: {df.columns.tolist()}")

    assume_tz = _detect_assume_tz(date_col)

    # Собираем raw сделки
    raw_trades = []
    fee_col = get_col('fee')
    row_side_values = []

    for _, row in df.iterrows():
        try:
            # Парсинг даты — naive datetime интерпретируем в assume_tz и переводим в UTC.
            entry_at = _normalize_to_utc_naive(row[date_col], assume_tz)
            if entry_at is None:
                continue

            # Парсинг направления
            raw_side = str(row[side_col]).lower()
            row_side_values.append(raw_side)
            if 'buy' in raw_side or 'long' in raw_side:
                direction = models.TradeDirection.LONG
            elif 'sell' in raw_side or 'short' in raw_side:
                direction = models.TradeDirection.SHORT
            else:
                continue # Пропускаем неизвестные типы (напр. Transfer)

            # Числовые поля
            price = float(row[price_col])
            quantity = float(row[qty_col])
            pnl = None
            if pnl_col and pd.notna(row.get(pnl_col)):
                try:
                    pnl = float(row[pnl_col])
                except Exception:
                    pnl = None
            
            # Комиссия
            commission = 0.0
            if fee_col and pd.notna(row.get(fee_col)):
                try:
                    commission = abs(float(row[fee_col]))
                except (ValueError, TypeError):
                    pass

            # Валюта: берём из колонки если есть, иначе оставляем None —
            # пусть на уровне раскладки в позицию подставится дефолт по аккаунту,
            # а не молча «RUB» для USD-инструментов.
            currency = None
            if currency_col and pd.notna(row.get(currency_col)):
                cur_raw = str(row[currency_col]).strip().upper()
                if cur_raw:
                    currency = cur_raw

            raw_trades.append({
                "symbol": str(row[symbol_col]).upper(),
                "direction": direction,
                "entry_price": price,
                "quantity": quantity,
                "entry_at": entry_at,
                "pnl": pnl,
                "commission": commission,
                "deal_sum": price * quantity,
                "swap": 0.0,
                "currency": currency,
                "notes": f"Imported from {filename}",
                "tags": ["Imported"]
            })
            
        except Exception as e:
            log.warning(f"Skipping row due to error: {e}")
            continue

    if not raw_trades:
        return trades
    
    # Сортируем по времени для корректной обработки
    raw_trades.sort(key=lambda x: x['entry_at'])

    def should_group_as_positions() -> bool:
        # Явные LONG/SHORT CSV обычно описывают уже позиции, а не отдельные execution rows.
        if any('long' in side or 'short' in side for side in row_side_values):
            return True

        # BUY/SELL CSV агрегируем только если видно усреднение/пирамидинг:
        # повторный вход тем же символом в том же направлении до закрытия.
        open_directions = {}
        for trade in raw_trades:
            symbol = trade['symbol']
            direction = trade['direction']
            current_direction = open_directions.get(symbol)

            if current_direction is None:
                open_directions[symbol] = direction
                continue

            if current_direction == direction:
                return True

            open_directions.pop(symbol, None)

        return False

    if should_group_as_positions():
        manager = TradeManager()
        for trade_data in raw_trades:
            manager.process_trade(trade_data)

        result = manager.finalize()

        fields_to_remove = ['deal_sum', '_total_cost', '_deal_sum']
        for trade in result:
            for field in fields_to_remove:
                trade.pop(field, None)
    else:
        result = []
        for trade in raw_trades:
            clean_trade = trade.copy()
            clean_trade.pop('deal_sum', None)
            clean_trade.setdefault('asset_name', None)
            clean_trade.setdefault('asset_type', 'Stock')
            clean_trade.setdefault('exit_at', None)
            clean_trade.setdefault('exit_price', None)
            clean_trade.setdefault('exit_reason', None)
            clean_trade.setdefault('entry_commission', clean_trade.get('commission', 0.0))
            clean_trade.setdefault('exit_commission', 0.0)
            clean_trade.setdefault('holding_time_minutes', None)
            # currency не подменяем «RUB» по умолчанию — None означает «брать из аккаунта»
            if clean_trade.get('currency') is None:
                clean_trade.pop('currency', None)
            clean_trade.setdefault('operations', [{
                "type": "entry",
                "time": clean_trade['entry_at'].strftime("%H:%M:%S") if isinstance(clean_trade.get('entry_at'), datetime) else str(clean_trade.get('entry_at')),
                "date": clean_trade['entry_at'].strftime("%d.%m.%Y") if isinstance(clean_trade.get('entry_at'), datetime) else str(clean_trade.get('entry_at')),
                "price": clean_trade.get('entry_price'),
                "qty": clean_trade.get('quantity'),
                "commission": clean_trade.get('commission', 0.0),
                "direction": clean_trade['direction'].value if hasattr(clean_trade.get('direction'), 'value') else str(clean_trade.get('direction')),
            }])
            if clean_trade.get('pnl') is not None:
                clean_trade['net_pnl'] = clean_trade['pnl'] - clean_trade['commission'] - clean_trade['swap']
            else:
                clean_trade['net_pnl'] = None
            result.append(clean_trade)
    
    # Сортируем по дате входа для корректного отображения в UI
    result.sort(key=lambda x: x.get('entry_at') or datetime.min)
    
    return result
